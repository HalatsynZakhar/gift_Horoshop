from __future__ import annotations

import io
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from horoshop_gifts import (
    CatalogIndex,
    GiftPlan,
    GiftRow,
    build_excel_template,
    build_registry_excel,
    import_results,
    load_settings,
    mutate_gifts,
    parse_excel_gifts,
    prepare_plan,
)


class HoroshopGiftsTests(unittest.TestCase):
    def excel_bytes(self, rows: list[tuple[object, ...]]) -> bytes:
        workbook = Workbook()
        for row in rows:
            workbook.active.append(row)
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_template_has_two_input_columns_and_guide(self) -> None:
        workbook = load_workbook(io.BytesIO(build_excel_template()), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Подарунки", "Інструкція"])
        self.assertEqual([cell.value for cell in next(workbook["Подарунки"].iter_rows(max_row=1))], ["Артикул основного товару", "Артикул подарунка"])
        self.assertIn("не стираються", workbook["Інструкція"]["A6"].value)
        workbook.close()

    def test_excel_parses_add_and_delete_marker(self) -> None:
        rows = parse_excel_gifts(self.excel_bytes([
            ("Артикул основного товару", "Артикул подарунка", "Видалити (Так)"),
            ("MAIN", "GIFT", ""),
            ("MAIN-2", "GIFT-2", "Так"),
        ]))
        self.assertEqual(rows[0], GiftRow("MAIN", "GIFT", 2))
        self.assertEqual(rows[1].action, "delete")

    def test_catalog_resolves_display_articles_and_preserves_gifts(self) -> None:
        catalog = CatalogIndex.from_raw([
            {"article": "MAIN-REAL", "article_for_display": "Main", "gifts": ["OLD-GIFT", {"page": {"id": 12}}]},
            {"article": "NEW-GIFT", "article_for_display": "new-gift"},
            {"article": "OLD-GIFT", "article_for_display": "Old-Gift"},
        ])
        plan = prepare_plan([GiftRow("Main", "NEW-GIFT", 2)], catalog)
        self.assertTrue(plan[0].ready)
        gifts, messages = mutate_gifts(catalog.by_article["MAIN-REAL"].gifts, plan)
        self.assertEqual(gifts, ["OLD-GIFT", {"page": {"id": 12}}, "NEW-GIFT"])
        self.assertIn("додано", messages[("MAIN-REAL", "NEW-GIFT")])

    def test_delete_mutation_removes_only_selected_gift(self) -> None:
        plans = [GiftPlan("Main", "Gift", "MAIN", "REMOVE", 2, "delete")]
        gifts, _ = mutate_gifts(("KEEP", "REMOVE", {"page": "Special"}), plans)
        self.assertEqual(gifts, ["KEEP", {"page": "Special"}])

    def test_registry_export_has_delete_marker(self) -> None:
        from horoshop_gifts import GiftAssociation
        workbook = load_workbook(io.BytesIO(build_registry_excel([GiftAssociation("Main", "Gift", "M", "G")])), read_only=True)
        self.assertEqual(workbook.active["C1"].value, "Видалити (Так)")
        self.assertIsNone(workbook.active["C2"].value)
        workbook.close()

    def test_catalog_import_result_uses_success_code(self) -> None:
        results = import_results({"status": "WARNING", "response": {"log": [{"article": "MAIN", "info": [{"code": 0, "message": "Товар оновлено"}]}]}})
        self.assertEqual(results["MAIN"], (True, "Товар оновлено"))

    def test_settings_repairs_unescaped_public_path(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            config = Path(directory) / "config.json"
            config.write_text(r'{"horoshop":{"domain":"https://shop.example.com"},"logging":{"public_log_path":"C:\ShareFiles\public","public_log_name":"gifts.log"}}', encoding="utf-8")
            settings = load_settings(config)
            repaired = config.read_text(encoding="utf-8")
        self.assertEqual(settings.public_log_file, Path(r"C:\ShareFiles\public") / "gifts.log")
        self.assertIn(r'"public_log_path":"C:\\ShareFiles\\public"', repaired)


if __name__ == "__main__":
    unittest.main()
