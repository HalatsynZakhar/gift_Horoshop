from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_EXPORT_LIMIT = 500
PUBLIC_LOG_PATH_JSON_PATTERN = re.compile(r'("public_log_path"\s*:\s*")(?P<value>[^"]*)(")')


class HoroshopGiftsError(RuntimeError):
    pass


@dataclass(frozen=True)
class Settings:
    domain: str
    host: str
    port: int
    batch_size: int
    request_timeout_seconds: int
    public_log_path: Path
    public_log_name: str

    @property
    def public_log_file(self) -> Path:
        return self.public_log_path / self.public_log_name


@dataclass(frozen=True)
class Credentials:
    login: str
    password: str
    token: str = ""


@dataclass(frozen=True)
class GiftRow:
    primary_display: str
    gift_display: str
    row_number: int
    action: str = "upsert"


@dataclass(frozen=True)
class CatalogProduct:
    article: str
    article_for_display: str
    gifts: tuple[Any, ...]


@dataclass(frozen=True)
class GiftPlan:
    primary_display: str
    gift_display: str
    primary_article: str
    gift_article: str
    row_number: int
    action: str = "upsert"
    error: str = ""

    @property
    def ready(self) -> bool:
        return not self.error


@dataclass(frozen=True)
class GiftAssociation:
    primary_display: str
    gift_display: str
    primary_article: str
    gift_article: str


def normalize(value: Any) -> str:
    return "" if value is None else str(value).strip()


def endpoint_url(domain: str, endpoint: str) -> str:
    return urljoin(f"{domain.rstrip('/')}/", endpoint.lstrip("/"))


def repair_public_log_path_json(config_text: str) -> str:
    def replace_path(match: re.Match[str]) -> str:
        value = re.sub(r"(?<!\\)\\(?!\\)", lambda _: "\\\\", match.group("value"))
        return f'{match.group(1)}{value}{match.group(3)}'

    return PUBLIC_LOG_PATH_JSON_PATTERN.sub(replace_path, config_text)


def load_settings(config_file: Path) -> Settings:
    config_text = config_file.read_text(encoding="utf-8-sig")
    try:
        raw = json.loads(config_text)
    except json.JSONDecodeError as original_error:
        repaired = repair_public_log_path_json(config_text)
        if repaired == config_text:
            raise original_error
        try:
            raw = json.loads(repaired)
        except json.JSONDecodeError:
            raise original_error
        try:
            config_file.write_text(repaired, encoding="utf-8")
        except OSError:
            pass
    if not isinstance(raw, dict):
        raise ValueError("config.json must contain an object.")

    horoshop = raw.get("horoshop") or {}
    server = raw.get("server") or {}
    logging_config = raw.get("logging") or {}
    if not isinstance(horoshop, dict) or not isinstance(server, dict) or not isinstance(logging_config, dict):
        raise ValueError("Sections horoshop, server and logging must contain objects.")
    domain = normalize(horoshop.get("domain"))
    if not domain:
        raise ValueError("Set horoshop.domain in config.json.")

    public_log_path = Path(normalize(logging_config.get("public_log_path", "logs")) or "logs")
    if not public_log_path.is_absolute():
        public_log_path = config_file.parent / public_log_path
    public_log_name = normalize(logging_config.get("public_log_name", "horoshop_gifts.log")) or "horoshop_gifts.log"
    if Path(public_log_name).name != public_log_name:
        raise ValueError("logging.public_log_name must be a file name without a path.")
    return Settings(
        domain=domain.rstrip("/"),
        host=normalize(server.get("host", "0.0.0.0")) or "0.0.0.0",
        port=max(1, min(65535, int(server.get("port", 8094)))),
        batch_size=max(1, int(horoshop.get("batch_size", 50))),
        request_timeout_seconds=max(1, int(horoshop.get("request_timeout_seconds", 60))),
        public_log_path=public_log_path,
        public_log_name=public_log_name,
    )


def parse_action(value: Any) -> str:
    marker = normalize(value).casefold()
    if marker in {"", "ні", "нет", "no", "0"}:
        return "upsert"
    if marker in {"так", "да", "yes", "1", "видалити", "удалить", "delete"}:
        return "delete"
    raise ValueError("у колонці «Видалити (Так)» вкажіть Так або залиште її порожньою.")


def parse_excel_gifts(data: bytes) -> list[GiftRow]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows: list[GiftRow] = []
        seen: set[tuple[str, str]] = set()
        headers = {"основний товар", "артикул основного товару", "primary article", "основной товар"}
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if not row or all(value is None for value in row[:3]):
                continue
            primary = normalize(row[0] if len(row) > 0 else "")
            if primary.casefold() in headers:
                continue
            gift = normalize(row[1] if len(row) > 1 else "")
            try:
                action = parse_action(row[2] if len(row) > 2 else "")
                if not primary:
                    raise ValueError("вкажіть артикул основного товару.")
                if not gift:
                    raise ValueError("вкажіть артикул подарунка.")
                key = (primary.casefold(), gift.casefold())
                if key in seen:
                    raise ValueError("цей зв'язок товару й подарунка повторюється.")
                seen.add(key)
            except ValueError as error:
                raise ValueError(f"Рядок {row_number}: {error}") from error
            rows.append(GiftRow(primary, gift, row_number, action))
        if not rows:
            raise ValueError("Excel не містить жодного зв'язку подарунка.")
        return rows
    finally:
        workbook.close()


def _style_header(worksheet: Any) -> None:
    fill = PatternFill("solid", fgColor="166534")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def build_excel_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Подарунки"
    worksheet.append(["Артикул основного товару", "Артикул подарунка"])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:B1"
    worksheet.column_dimensions["A"].width = 36
    worksheet.column_dimensions["B"].width = 36
    _style_header(worksheet)
    for row in range(2, 102):
        worksheet.cell(row=row, column=1).number_format = "@"
        worksheet.cell(row=row, column=2).number_format = "@"

    guide = workbook.create_sheet("Інструкція")
    guide.column_dimensions["A"].width = 105
    guide["A1"] = "Шаблон для призначення подарунків"
    guide["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="166534")
    guide.merge_cells("A1:B1")
    instructions = [
        "Заповнюйте лист «Подарунки». Кожен рядок додає один подарунок до одного товару.",
        "В обох колонках вказуйте article_for_display — артикул, що відображається на сайті.",
        "Сервіс знаходить внутрішні артикули в експорті каталогу та передає їх Хорошопу.",
        "Наявні подарунки основного товару не стираються: додається лише новий зв'язок.",
        "Для масового видалення вивантажте поточний список: у ньому є колонка «Видалити (Так)».",
    ]
    for row, instruction in enumerate(instructions, start=3):
        guide.cell(row=row, column=1, value=instruction).alignment = Alignment(wrap_text=True, vertical="top")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def build_registry_excel(associations: list[GiftAssociation]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Подарунки"
    worksheet.append(["Артикул основного товару", "Артикул подарунка", "Видалити (Так)"])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:C1"
    worksheet.column_dimensions["A"].width = 36
    worksheet.column_dimensions["B"].width = 36
    worksheet.column_dimensions["C"].width = 18
    _style_header(worksheet)
    for association in associations:
        worksheet.append([association.primary_display, association.gift_display, ""])
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=1).number_format = "@"
        worksheet.cell(row=row, column=2).number_format = "@"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def gift_article(value: Any) -> str:
    if isinstance(value, dict):
        return normalize(value.get("article"))
    return normalize(value)


class CatalogIndex:
    def __init__(self, products: list[CatalogProduct]) -> None:
        self.products = products
        self.by_article = {product.article: product for product in products}
        self.by_display: dict[str, list[CatalogProduct]] = {}
        self.by_display_folded: dict[str, list[CatalogProduct]] = {}
        for product in products:
            if product.article_for_display:
                self.by_display.setdefault(product.article_for_display, []).append(product)
                self.by_display_folded.setdefault(product.article_for_display.casefold(), []).append(product)

    @classmethod
    def from_raw(cls, raw_products: list[dict[str, Any]]) -> "CatalogIndex":
        products: list[CatalogProduct] = []
        for raw in raw_products:
            article = normalize(raw.get("article")) if isinstance(raw, dict) else ""
            if not article:
                continue
            gifts = raw.get("gifts", [])
            products.append(CatalogProduct(
                article=article,
                article_for_display=normalize(raw.get("article_for_display")),
                gifts=tuple(gifts) if isinstance(gifts, list) else (),
            ))
        return cls(products)

    def resolve_display(self, value: str) -> tuple[CatalogProduct | None, str]:
        exact = self.by_display.get(value, [])
        if len(exact) == 1:
            return exact[0], ""
        if len(exact) > 1:
            return None, f"Артикул відображення '{value}' не є унікальним."
        insensitive = self.by_display_folded.get(value.casefold(), [])
        if len(insensitive) == 1:
            return insensitive[0], ""
        if len(insensitive) > 1:
            return None, f"Артикул відображення '{value}' не є унікальним."
        return None, f"Артикул відображення '{value}' не знайдений у каталозі."

    def associations(self) -> list[GiftAssociation]:
        pairs: list[GiftAssociation] = []
        for primary in self.products:
            for raw_gift in primary.gifts:
                article = gift_article(raw_gift)
                if not article:
                    continue
                gift = self.by_article.get(article)
                pairs.append(GiftAssociation(
                    primary_display=primary.article_for_display or primary.article,
                    gift_display=(gift.article_for_display or gift.article) if gift else article,
                    primary_article=primary.article,
                    gift_article=article,
                ))
        return sorted(pairs, key=lambda pair: (pair.primary_display.casefold(), pair.gift_display.casefold()))


def prepare_plan(rows: list[GiftRow], catalog: CatalogIndex) -> list[GiftPlan]:
    plan: list[GiftPlan] = []
    for row in rows:
        primary, error = catalog.resolve_display(row.primary_display)
        if error:
            plan.append(GiftPlan(row.primary_display, row.gift_display, "", "", row.row_number, row.action, error))
            continue
        gift, error = catalog.resolve_display(row.gift_display)
        if error:
            plan.append(GiftPlan(row.primary_display, row.gift_display, primary.article if primary else "", "", row.row_number, row.action, error))
            continue
        if primary is not None and gift is not None and primary.article == gift.article:
            error = "Основний товар і подарунок не можуть бути одним товаром."
        plan.append(GiftPlan(
            row.primary_display,
            row.gift_display,
            primary.article if primary else "",
            gift.article if gift else "",
            row.row_number,
            row.action,
            error,
        ))
    return plan


def mutate_gifts(existing: tuple[Any, ...], plans: list[GiftPlan]) -> tuple[list[Any], dict[tuple[str, str], str]]:
    gifts = list(existing)
    messages: dict[tuple[str, str], str] = {}
    for plan in plans:
        key = plan.gift_article.casefold()
        positions = [index for index, value in enumerate(gifts) if gift_article(value).casefold() == key]
        row_key = (plan.primary_article, plan.gift_article)
        if plan.action == "delete":
            if positions:
                gifts = [value for value in gifts if gift_article(value).casefold() != key]
                messages[row_key] = "Подарунок буде видалено."
            else:
                messages[row_key] = "Цей подарунок не був призначений товару."
        elif positions:
            messages[row_key] = "Подарунок уже призначений товару."
        else:
            gifts.append(plan.gift_article)
            messages[row_key] = "Подарунок буде додано."
    return gifts, messages


class HoroshopClient:
    def __init__(self, settings: Settings, credentials: Credentials) -> None:
        self.settings = settings
        self.credentials = credentials
        self.session = requests.Session()
        self._token = credentials.token

    def _post(self, endpoint: str, payload: dict[str, Any]) -> dict[str, Any]:
        try:
            response = self.session.post(endpoint_url(self.settings.domain, endpoint), json=payload, timeout=self.settings.request_timeout_seconds)
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as error:
            raise HoroshopGiftsError(f"Horoshop API request failed: {error}") from error
        except ValueError as error:
            raise HoroshopGiftsError("Horoshop API returned a non-JSON response.") from error
        if not isinstance(data, dict):
            raise HoroshopGiftsError("Horoshop API returned an invalid JSON response.")
        if str(data.get("status", "")).upper() in {"ERROR", "EXCEPTION"}:
            raise HoroshopGiftsError(str(data))
        return data

    def token(self) -> str:
        if self._token:
            return self._token
        if not self.credentials.login or not self.credentials.password:
            raise HoroshopGiftsError("Вкажіть логін і пароль API або чинний токен.")
        response = self._post("/api/auth/", {"login": self.credentials.login, "password": self.credentials.password})
        token = response.get("response", {}).get("token")
        if not token:
            raise HoroshopGiftsError("Хорошоп не повернув токен авторизації.")
        self._token = str(token)
        return self._token

    def export_catalog(self) -> list[dict[str, Any]]:
        offset = 0
        products: list[dict[str, Any]] = []
        while True:
            response = self._post("/api/catalog/export/", {
                "token": self.token(),
                "offset": offset,
                "limit": DEFAULT_EXPORT_LIMIT,
                "includedParams": ["article_for_display", "gifts"],
            })
            nested = response.get("response")
            page = nested.get("products") if isinstance(nested, dict) else response.get("products")
            if not isinstance(page, list):
                raise HoroshopGiftsError("Експорт каталогу не містить товарів.")
            products.extend(item for item in page if isinstance(item, dict))
            if len(page) < DEFAULT_EXPORT_LIMIT:
                return products
            offset += DEFAULT_EXPORT_LIMIT

    def import_products(self, products: list[dict[str, Any]]) -> dict[str, Any]:
        if not products:
            return {"status": "OK", "response": {"log": []}}
        return self._post("/api/catalog/import/", {"token": self.token(), "products": products})


def import_results(response: dict[str, Any]) -> dict[str, tuple[bool, str]]:
    status = str(response.get("status", "")).upper()
    nested = response.get("response")
    log_items = nested.get("log", []) if isinstance(nested, dict) else []
    results: dict[str, tuple[bool, str]] = {}
    if not isinstance(log_items, list):
        return results
    for entry in log_items:
        if not isinstance(entry, dict):
            continue
        article = normalize(entry.get("article"))
        info = entry.get("info", [])
        codes: list[Any] = []
        messages: list[str] = []
        if isinstance(info, list):
            for item in info:
                if isinstance(item, dict):
                    codes.append(item.get("code"))
                    messages.append(normalize(item.get("message")))
        results[article] = (0 in codes or (status == "OK" and not codes), "; ".join(message for message in messages if message))
    return results
